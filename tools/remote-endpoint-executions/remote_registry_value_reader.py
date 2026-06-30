#!/usr/bin/env python3
"""
remote_registry_value_reader.py

Purpose:
    Read one registry value from multiple authorized remote Windows endpoints.

Created by:
    Avraham Makovsky

License:
    MIT

Why it exists:
    In infrastructure and support work, it is often useful to check the same
    registry value across many workstations without opening each machine
    manually.

How it works:
    Uses Python's winreg remote registry access:
        winreg.ConnectRegistry(host, hive)

    This is not WMI. The target machines must allow remote registry access,
    and the running user must have permission to read the requested key/value.

Typical examples:
    python remote_registry_value_reader.py ^
        --registry-path "SOFTWARE\\Vendor\\Product" ^
        --value-name "Version" ^
        --notepad

    python remote_registry_value_reader.py ^
        --registry-path "SOFTWARE\\Vendor\\Product" ^
        --value-name "Version" ^
        --hosts hosts.txt ^
        --output results.csv

    python remote_registry_value_reader.py ^
        --hive HKLM ^
        --registry-path "SOFTWARE\\Vendor\\Product" ^
        --value-name "Version" ^
        --hosts hosts.txt ^
        --excel results.xlsx

Dependencies:
    CSV export uses only the Python standard library.
    Excel export is optional and requires openpyxl:
        pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

try:
    import winreg
except ImportError:
    print("Error: winreg is available only on Windows.", file=sys.stderr)
    raise SystemExit(1)


DEFAULT_MAX_WORKERS = 16

RESULT_COLUMNS = [
    "Host",
    "Hive",
    "RegistryPath",
    "ValueName",
    "ValueData",
    "ValueType",
    "Status",
    "Details",
]

HIVE_MAP = {
    "HKLM": winreg.HKEY_LOCAL_MACHINE,
    "HKEY_LOCAL_MACHINE": winreg.HKEY_LOCAL_MACHINE,
    "HKU": winreg.HKEY_USERS,
    "HKEY_USERS": winreg.HKEY_USERS,
}

TYPE_MAP = {
    winreg.REG_SZ: "REG_SZ",
    winreg.REG_EXPAND_SZ: "REG_EXPAND_SZ",
    winreg.REG_MULTI_SZ: "REG_MULTI_SZ",
    winreg.REG_DWORD: "REG_DWORD",
    winreg.REG_QWORD: "REG_QWORD",
    winreg.REG_BINARY: "REG_BINARY",
    winreg.REG_NONE: "REG_NONE",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read one registry value from multiple authorized remote Windows endpoints."
    )

    parser.add_argument(
        "--hive",
        default="HKLM",
        choices=sorted(HIVE_MAP.keys()),
        help="Registry hive. Default: HKLM.",
    )
    parser.add_argument(
        "--registry-path",
        "-p",
        required=True,
        help=r'Registry path under the selected hive. Example: "SOFTWARE\Vendor\Product"',
    )
    parser.add_argument(
        "--value-name",
        "-v",
        required=True,
        help='Registry value name to read. Use "(Default)" for the default value.',
    )
    parser.add_argument(
        "--hosts",
        help="Text file containing hostnames. Lines starting with # or ; are ignored.",
    )
    parser.add_argument(
        "--notepad",
        action="store_true",
        help="Open Notepad to paste hostnames. Windows only.",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="CSV output path. Default: timestamped CSV in the current folder.",
    )
    parser.add_argument(
        "--excel",
        help="Optional Excel output path. Requires openpyxl.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_MAX_WORKERS,
        help=f"Maximum parallel workers. Default: {DEFAULT_MAX_WORKERS}.",
    )
    parser.add_argument(
        "--keep-order",
        action="store_true",
        help="Keep input host order in the final output. Default: sort by hostname.",
    )

    return parser.parse_args()


def require_windows() -> None:
    if os.name != "nt":
        raise RuntimeError("This tool is intended to run on Windows.")


def normalize_registry_path(path: str) -> str:
    cleaned = path.strip().strip("\\/")
    if not cleaned:
        raise ValueError("Registry path cannot be empty.")
    return cleaned.replace("/", "\\")


def normalize_value_name(value_name: str) -> str:
    if value_name.strip().lower() in {"(default)", "default", ""}:
        return ""
    return value_name.strip()


def parse_hosts(text: str) -> List[str]:
    """Parse hostnames from lines, spaces, or commas while preserving order."""
    hosts: List[str] = []
    seen = set()

    for raw_line in text.splitlines():
        line = raw_line.strip()

        if not line or line.startswith("#") or line.startswith(";"):
            continue

        for item in line.replace(",", " ").split():
            host = item.strip()

            if not host or host.startswith("#") or host.startswith(";"):
                continue

            key = host.lower()

            if key not in seen:
                seen.add(key)
                hosts.append(host)

    return hosts


def load_hosts_from_file(path: str) -> List[str]:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    return parse_hosts(text)


def load_hosts_from_notepad() -> List[str]:
    """Collect hostnames through Notepad for quick ad-hoc use."""
    require_windows()
    temp_path = None

    try:
        fd, temp_path = tempfile.mkstemp(prefix="remote_registry_hosts_", suffix=".txt", text=True)
        os.close(fd)

        Path(temp_path).write_text(
            "# Paste one hostname per line.\n"
            "# Lines starting with # or ; are ignored.\n"
            "LAB-PC-001\n"
            "LAB-PC-002\n",
            encoding="utf-8",
        )

        subprocess.run(["notepad.exe", temp_path], check=False)
        return load_hosts_from_file(temp_path)

    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except OSError:
                pass


def registry_type_to_string(reg_type: int) -> str:
    return TYPE_MAP.get(reg_type, f"REG_TYPE_{reg_type}")


def value_to_string(value: Any, reg_type: int) -> str:
    """Convert registry data into a safe, readable string for CSV/console output."""
    if value is None:
        return ""

    if reg_type == winreg.REG_BINARY:
        return bytes(value).hex(" ").upper()

    if reg_type == winreg.REG_MULTI_SZ:
        if isinstance(value, (list, tuple)):
            return "; ".join(str(item) for item in value)
        return str(value)

    if reg_type in {winreg.REG_DWORD, winreg.REG_QWORD}:
        return str(value)

    return str(value)


def make_base_result(host: str, hive_name: str, registry_path: str, value_name: str) -> Dict[str, str]:
    return {
        "Host": host,
        "Hive": hive_name,
        "RegistryPath": registry_path,
        "ValueName": value_name or "(Default)",
        "ValueData": "",
        "ValueType": "",
        "Status": "Failed",
        "Details": "",
    }


def read_remote_registry_value(
    host: str,
    hive_name: str,
    hive_handle: int,
    registry_path: str,
    value_name: str,
) -> Dict[str, str]:
    """Read one registry value from one remote host."""
    result = make_base_result(host, hive_name, registry_path, value_name)
    remote_registry = None
    key = None

    try:
        remote_registry = winreg.ConnectRegistry(rf"\\{host}", hive_handle)
        key = winreg.OpenKey(remote_registry, registry_path, 0, winreg.KEY_READ)
        value_data, value_type = winreg.QueryValueEx(key, value_name)

        result["ValueData"] = value_to_string(value_data, value_type)
        result["ValueType"] = registry_type_to_string(value_type)
        result["Status"] = "OK"
        result["Details"] = ""

    except FileNotFoundError as exc:
        result["Status"] = "NotFound"
        result["Details"] = str(exc)

    except PermissionError as exc:
        result["Status"] = "AccessDenied"
        result["Details"] = str(exc)

    except OSError as exc:
        result["Status"] = "Failed"
        result["Details"] = str(exc)

    except Exception as exc:
        result["Status"] = "Failed"
        result["Details"] = f"{type(exc).__name__}: {exc}"

    finally:
        if key is not None:
            try:
                winreg.CloseKey(key)
            except OSError:
                pass

        if remote_registry is not None:
            try:
                winreg.CloseKey(remote_registry)
            except OSError:
                pass

    return result


def default_csv_path() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"remote_registry_value_reader_{stamp}.csv"


def write_csv(path: str, rows: Iterable[Dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in RESULT_COLUMNS})


def write_excel(path: str, rows: List[Dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl. Install with: pip install openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Registry Results"
    worksheet.append(RESULT_COLUMNS)

    for row in rows:
        worksheet.append([row.get(column, "") for column in RESULT_COLUMNS])

    for column_index, column_name in enumerate(RESULT_COLUMNS, start=1):
        max_len = len(column_name)
        column_letter = get_column_letter(column_index)

        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_len + 2, 80)

    if rows:
        end_cell = f"{get_column_letter(len(RESULT_COLUMNS))}{len(rows) + 1}"
        table = Table(displayName="RegistryResults", ref=f"A1:{end_cell}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    workbook.save(path)


def print_progress(index: int, total: int, row: Dict[str, str]) -> None:
    host = row["Host"]
    status = row["Status"]

    if status == "OK":
        print(f"[{index}/{total}] {host:<24} OK           {row['ValueData']}")
    elif status == "NotFound":
        print(f"[{index}/{total}] {host:<24} Not found")
    elif status == "AccessDenied":
        print(f"[{index}/{total}] {host:<24} Access denied")
    else:
        print(f"[{index}/{total}] {host:<24} Failed       {row['Details'][:100]}")


def print_summary(rows: List[Dict[str, str]]) -> None:
    ok = sum(1 for row in rows if row["Status"] == "OK")
    not_found = sum(1 for row in rows if row["Status"] == "NotFound")
    access_denied = sum(1 for row in rows if row["Status"] == "AccessDenied")
    failed = sum(1 for row in rows if row["Status"] == "Failed")

    print()
    print("Summary")
    print("-------")
    print(f"Hosts checked: {len(rows)}")
    print(f"OK: {ok}")
    print(f"Not found: {not_found}")
    print(f"Access denied: {access_denied}")
    print(f"Failed: {failed}")


def get_hosts(args: argparse.Namespace) -> List[str]:
    if args.hosts:
        return load_hosts_from_file(args.hosts)

    if args.notepad or sys.stdin.isatty():
        return load_hosts_from_notepad()

    return parse_hosts(sys.stdin.read())


def main() -> int:
    try:
        require_windows()

        args = parse_args()
        hive_name = args.hive.upper()
        hive_handle = HIVE_MAP[hive_name]
        registry_path = normalize_registry_path(args.registry_path)
        value_name = normalize_value_name(args.value_name)
        hosts = get_hosts(args)

        if not hosts:
            print("No hosts provided.", file=sys.stderr)
            return 1

        workers = max(1, min(args.workers, 64))

        print("Remote Registry Value Reader")
        print("----------------------------")
        print(f"Hive:  {hive_name}")
        print(f"Path:  {registry_path}")
        print(f"Value: {value_name or '(Default)'}")
        print(f"Hosts: {len(hosts)}")
        print()

        rows_by_host: Dict[str, Dict[str, str]] = {}

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(
                    read_remote_registry_value,
                    host,
                    hive_name,
                    hive_handle,
                    registry_path,
                    value_name,
                ): host
                for host in hosts
            }

            for index, future in enumerate(as_completed(future_map), start=1):
                host = future_map[future]
                row = future.result()
                rows_by_host[host.lower()] = row
                print_progress(index, len(hosts), row)

        if args.keep_order:
            rows = [rows_by_host[host.lower()] for host in hosts]
        else:
            rows = sorted(rows_by_host.values(), key=lambda row: row["Host"].lower())

        print_summary(rows)

        csv_path = args.output or default_csv_path()
        write_csv(csv_path, rows)
        print()
        print(f"CSV saved: {csv_path}")

        if args.excel:
            write_excel(args.excel, rows)
            print(f"Excel saved: {args.excel}")

        return 0

    except KeyboardInterrupt:
        print("\nCanceled.")
        return 130

    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
